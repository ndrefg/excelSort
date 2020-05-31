from openpyxl import Workbook
from openpyxl import load_workbook

# -----Start Config----------------#

KEYS = ["Banana", "Cenas", "Cherries"] # Lista com os valores a procurar


ITERATOR_START = 1 # row onde começa a escrever no resultado


UNDEF_START = 1 # row onde mete os valores nao definidos


INPUT_FILE = 'produceSales.xlsx' # Ficheiro original a ler


OUTPUT_FILE = 'file2.xlsx' # Ficheiro a escrever e que vai ser criado se nao existir

# Coluna onde estão os dados a procurar
KEY_COL = 0

# -----End Config------------------#


try:

    # if ITERATOR_START >= UNDEF_START:
    #    raise Exception('iterator is greater than undef. This will result in data being overwritten in the output file')
    iterator = ITERATOR_START # define o interator com o valor da row onde começa a escrever no resultado
    undef = UNDEF_START # reduz a variavel UNDEF_START definida em cima para undef

    wb = load_workbook(filename=INPUT_FILE) # abre o ficheiro de origem
    sheets = wb.sheetnames # Abre as folhas do ficheiro de origem

    newWb = Workbook() # Cria e abre ou só abre o ficheiro de destino
    newSheet = newWb.sheetnames # Abre as folhas do ficheiro de destino
    newWs = newWb[newSheet[0]] # Cria uma nova folha no ficheiro de destino
    ws3 = newWb.create_sheet(title="Data")

    ws = wb[sheets[0]] # Chama a folha 0 do ficheiro de origem de "ws"

    for i in KEYS: # cicla pelo numero de KEYS que estiverem na lista a procurar
        for row in ws.iter_rows(): # itera pelas rows todas

            if row[KEY_COL].value == i: # Se o valor na coluna a procurar da row for igual ao valor da lista
                for cell in row: # Para o valor na celula da row faz
                    newWs.cell(row=iterator, column=cell.column).value = cell.value #copia valores para ficheiro de destino
                iterator += 1 # adiciona 1 ao iterador para passar para a proxima row
                # if iterator >= UNDEF_START: # se o iterador for maior que 50
                #     raise Exception('Iterator has reached range reserved for undefined rows. Iterator: ' + str(
                #         iterator) + ' Undef_Start: ' + str(UNDEF_START))

    for row in ws.iter_rows(): # itera pelas rows todas
        if row[KEY_COL].value not in KEYS: # se o valor na coluna a procurar da row nao estiver nas lista
            for cell in row: # percorre as cell em pelas row
                ws3.cell(row=undef, column=cell.column).value = cell.value # copia os valores para o ficheiro de destino
            undef += 1 # adiciona 1 ao iterador para passar para a proxima row
    newWb.save(filename=OUTPUT_FILE) # guarda o ficheiro de destino

except FileNotFoundError as e:
    print(
        "ERROR: Unable to open file. File may be outside of application directory")  # Does not exist OR no read permissions
except PermissionError as e:
    print("ERROR: Output file is open. Please close it and try again")
except Exception as e:
    print('ERROR: ', e)
